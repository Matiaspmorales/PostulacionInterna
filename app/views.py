from django.db import transaction
from django.db.models import Q
from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from . models import  PostulacionInternos, PostulanteInterno, CompetenciaEspecifica, CompetenciaTransversal
from django.utils import timezone
from django.utils.encoding import smart_str
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from django.contrib.auth.decorators import user_passes_test
from app.forms import PostulacionInternosForm, PostulanteInternoForm, CompetenciaEspecificaFormSet, CompetenciaTransversalFormSet
from .utils import eliminar_archivos_existentes_internos
# Create your views here.

def index(request):
    return render(request, 'index.html')


def iniciarSesion(request):
    if request.method == "POST":
        username = request.POST.get('username', '').strip()
        contraseña = request.POST.get('password', '').strip()

        user = authenticate(request, username=username, password=contraseña)
        if user:
            login(request, user)
            messages.success(
                request, f"Bienvenido/a {user.username}", extra_tags="inicio_exitoso")
            return redirect('postulacion:index')
        else:
            messages.error(
                request, "Usuario o contraseña incorrectos.", extra_tags="inicio_fallido")
    return render(request, "inicio_sesion.html")


def cerrarSesion(request):
    logout(request)
    return redirect('postulacion:index')

def error_403(request, exception):
    return render(request, 'errores/403.html', status=403)

def error_404(request, exception):
    return render(request, 'errores/404.html', status=404)


def es_superuser(user):
    return user.is_superuser

def es_superuser_o_staff(user):
    return user.is_staff or user.is_superuser






@user_passes_test(es_superuser_o_staff, login_url='postulacion:acceso_denegado')   
def crear_postulacion_interna(request):
    if request.method == 'POST':
        form = PostulacionInternosForm(request.POST, request.FILES)
        formset_transversales = CompetenciaTransversalFormSet(request.POST, prefix='trans')
        formset_especificas = CompetenciaEspecificaFormSet(request.POST, prefix='esp')

        if form.is_valid() and formset_transversales.is_valid() and formset_especificas.is_valid():
            postulacion = form.save()
            for fs in formset_transversales.save(commit=False):
                fs.postulacion = postulacion
                fs.save()
            for fs in formset_especificas.save(commit=False):
                fs.postulacion = postulacion
                fs.save()
            messages.success(request, "Convocatoria creada correctamente.", extra_tags="convocatoria_exito")
        else:
            print("Errores en el formulario:")
            print("Formulario principal:", form.errors.as_json())
            print("Formset transversales:", formset_transversales.errors)
            print("Formset específicas:", formset_especificas.errors)
            messages.error(request, "Corrige los errores del formulario.", extra_tags="convocatoria_error")
    else:
        form = PostulacionInternosForm()
        formset_transversales = CompetenciaTransversalFormSet(queryset=CompetenciaTransversal.objects.none(), prefix='trans')
        formset_especificas = CompetenciaEspecificaFormSet(queryset=CompetenciaEspecifica.objects.none(), prefix='esp')

    return render(request, 'internos/formularios/form_crear_postulacion_interno.html', {
        'form': form,
        'formset_transversales': formset_transversales,
        'formset_especificas': formset_especificas,
    })

@user_passes_test(es_superuser_o_staff, login_url='postulacion:acceso_denegado')
def editar_postulacion_interna(request, postulacion_id):
    postulacion = get_object_or_404(PostulacionInternos, id=postulacion_id)

    if request.method == 'POST':
        form = PostulacionInternosForm(request.POST, request.FILES, instance=postulacion)
        formset_transversales = CompetenciaTransversalFormSet(
            request.POST, prefix='trans',
            queryset=CompetenciaTransversal.objects.filter(postulacion=postulacion)
        )
        formset_especificas = CompetenciaEspecificaFormSet(
            request.POST, prefix='esp',
            queryset=CompetenciaEspecifica.objects.filter(postulacion=postulacion)
        )

        if form.is_valid() and formset_transversales.is_valid() and formset_especificas.is_valid():
            form.save()

            for form_trans in formset_transversales:
                if form_trans.cleaned_data.get('DELETE') and form_trans.instance.pk:
                    form_trans.instance.delete()
                elif not form_trans.cleaned_data.get('DELETE'):
                    instancia = form_trans.save(commit=False)
                    instancia.postulacion = postulacion
                    instancia.save()

            for form_esp in formset_especificas:
                if form_esp.cleaned_data.get('DELETE') and form_esp.instance.pk:
                    form_esp.instance.delete()
                elif not form_esp.cleaned_data.get('DELETE'):
                    instancia = form_esp.save(commit=False)
                    instancia.postulacion = postulacion
                    instancia.save()

            messages.success(request, "Convocatoria actualizada correctamente.", extra_tags="convocatoria_exito")
        else:
            messages.error(request, "Corrige los errores del formulario.", extra_tags="convocatoria_error")
    else:
        form = PostulacionInternosForm(instance=postulacion)
        formset_transversales = CompetenciaTransversalFormSet(
            queryset=CompetenciaTransversal.objects.filter(postulacion=postulacion),
            prefix='trans'
        )
        formset_especificas = CompetenciaEspecificaFormSet(
            queryset=CompetenciaEspecifica.objects.filter(postulacion=postulacion),
            prefix='esp'
        )

    return render(request, 'internos/formularios/form_editar_postulacion_interno.html', {
        'form': form,
        'formset_transversales': formset_transversales,
        'formset_especificas': formset_especificas,
        'postulacion': postulacion,
    })


def postulacion_internas(request):
    postulaciones = PostulacionInternos.objects.all()
    hoy = timezone.localdate()

    return render(request, "internos/internos.html", {
        'postulaciones': postulaciones,
        'hoy': hoy
    })


def ingreso_postulante_interno(request, postulacion_id):
    postulacion = get_object_or_404(PostulacionInternos, id=postulacion_id)
    hoy = timezone.localdate()

    if postulacion.fecha_inicio > hoy or postulacion.fecha_termino < hoy:
        messages.error(request, "La postulación a esta convocatoria está cerrada.")
        return redirect("postulacion:index")

    competencias_trans = CompetenciaTransversal.objects.filter(postulacion=postulacion)
    competencias_esp = CompetenciaEspecifica.objects.filter(postulacion=postulacion)

    if request.method == "POST":
        form = PostulanteInternoForm(request.POST, request.FILES)
        rut = request.POST.get("rut")
        existente = PostulanteInterno.objects.filter(rut=rut, postulacion=postulacion).first()

        if form.is_valid():
            try:
                with transaction.atomic():
                    if existente:
                        eliminar_archivos_existentes_internos(existente)
                        existente.delete()

                    nuevo = form.save(commit=False)
                    nuevo.postulacion = postulacion
                    nuevo.save()

                    messages.success(
                        request,
                        "Postulación enviada correctamente. Al Aceptar será redirigido nuevamente al inicio.",
                        extra_tags="postulacion_exito"
                    )
                    form = PostulanteInternoForm()

            except Exception as e:
                print("Error al guardar:", e)
                messages.error(
                    request,
                    "Error al guardar la postulación.",
                    extra_tags="postulacion_error"
                )

        else:
            messages.error(
                request,
                "Formulario inválido. Revise los campos.",
                extra_tags="postulacion_error"
            )

    else:
        form = PostulanteInternoForm()

    return render(request, 'internos/formularios/form_postulacion_interno.html', {
        'form': form,
        'postulacion': postulacion,
        'competencias_trans': competencias_trans,
        'competencias_esp': competencias_esp
    })


@user_passes_test(es_superuser, login_url='postulacion:acceso_denegado')
def ver_postulantes_internos(request, id):
    postulacion = get_object_or_404(PostulacionInternos, id=id)

    search_query = request.GET.get('q', '').strip()
    postulantes = PostulanteInterno.objects.filter(postulacion=postulacion)

    if search_query:
        postulantes = postulantes.filter(
            Q(nombre_completo__icontains=search_query) |
            Q(rut__icontains=search_query)
        )

    return render(request, 'internos/tablas_formularios/tabla_postulacion_interna.html', {
        'postulacion': postulacion,
        'postulantes': postulantes,
        'search_query': search_query,
    })



@user_passes_test(es_superuser, login_url='postulacion:acceso_denegado')
def detalle_postulante_interno(request, id):
    postulante = get_object_or_404(PostulanteInterno, id=id)

    return render(request, 'internos/detalle/detalle_postulante_interno.html', {
        'postulante': postulante
    })


@user_passes_test(es_superuser, login_url='postulacion:acceso_denegado')
def exportar_postulantes_internos(request, postulacion_id):
    postulacion = get_object_or_404(PostulacionInternos, id=postulacion_id)
    postulantes = PostulanteInterno.objects.filter(postulacion=postulacion)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Postulantes - {postulacion.nombre_cargo[:25]}"

    encabezados = [
        "Fecha Creación", "RUT", "Nombre Completo", "Género", "Correo Electrónico",
        "Comuna", "Teléfono", "Años Experiencia", "Discapacidad",
        "¿Posee Título?", "¿Posee Cert. Laboral?", "¿Presenta Cursos?",
        "Cédula Identidad", "Currículum", "Certificado Título", "Certificado Laboral",
        "Cursos/Capacitaciones", "Cert. Discapacidad"
    ]

    # 1) Agregar encabezados
    ws.append(encabezados)

    # 2) Agregar filas de datos con URLs de documentos
    for postulante in postulantes:
        fila = [
            timezone.localtime(postulante.fecha_creacion).strftime("%Y-%m-%d %H:%M:%S"),
            postulante.rut,
            postulante.nombre_completo,
            postulante.genero,
            postulante.correo_electronico,
            postulante.comuna_residencia,
            postulante.telefono_contacto,
            postulante.años_experiencia,
            postulante.discapacidad,
            postulante.posee_certificado_titulo,
            postulante.posee_certificado_laboral,
            postulante.presenta_certificado_cursos,
        ]

        documentos = [
            postulante.cedula_identidad,
            postulante.curriculum,
            postulante.certificado_titulo,
            postulante.certificado_laboral,
            postulante.cursos_capacitaciones,
            postulante.certificado_discapacidad,
        ]

        urls_docs = [request.build_absolute_uri(d.url) if d else "" for d in documentos]
        fila_completa = fila + urls_docs
        ws.append(fila_completa)

        row = ws.max_row
        start_col = len(fila) + 1

        for idx, url in enumerate(urls_docs):
            col = start_col + idx
            cell = ws.cell(row=row, column=col)
            if url:
                cell.hyperlink = url
                cell.value = "Ver documento"
                cell.font = Font(color="0000FF", underline="single")

    # 3) Ajustar ancho de columnas
    total_columnas = len(encabezados)
    for idx in range(1, total_columnas + 1):
        letra = get_column_letter(idx)
        max_len = max((len(str(celda.value or "")) for celda in ws[letra]), default=0)
        ws.column_dimensions[letra].width = max_len + 2

    # 4) Aplicar borde fino a TODO el rango de celdas usadas
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000"),
    )
    max_row = ws.max_row
    for fila in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=total_columnas):
        for celda in fila:
            celda.border = thin_border

    # 5) Activar filtro automático en la fila de encabezados (fila 1)
    ultima_letra = get_column_letter(total_columnas)
    ws.auto_filter.ref = f"A1:{ultima_letra}1"

    # 6) Preparar respuesta HTTP con el archivo Excel
    nombre_archivo = f"postulantes_internos_{postulacion.nombre_cargo[:25].replace(' ', '_')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename={smart_str(nombre_archivo)}"
    wb.save(response)
    return response